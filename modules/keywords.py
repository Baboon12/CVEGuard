import openpyxl

from .cve_details import BASE_URL

def format_value(cell: str) -> str:
    cell_value = cell.value.lower()
    cell_value_split = cell_value.split(' ')
    if len(cell_value_split) > 1:
        cell_value = "+".join(cell_value_split)

    return cell_value


def get_keywords(keyword_file_path: str) -> list[str]:
    xl_keywords = openpyxl.load_workbook(keyword_file_path)
    sheet = xl_keywords.active
    max_rows = sheet.max_row

    keywords = []

    for i in range(1, max_rows + 1):
        cell = sheet.cell(row = i, column = 1)
        cell_value = format_value(cell)
        keywords.append(cell_value)

    return keywords


def create_urls(keywords: list[str]) -> list[str]:
    urls = []
    urls = [BASE_URL + keyword for keyword in keywords]

    return urls
